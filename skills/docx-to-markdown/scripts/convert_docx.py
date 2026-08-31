#!/usr/bin/env python3
"""
将docx文档转换为markdown格式，并提取所有图片到assets文件夹
支持将嵌入的Excel表格转换为Markdown表格
"""

import hashlib
import json
import logging
import math
import os
import sys
import zipfile
import re
import io
import unicodedata
from html import unescape
from html.parser import HTMLParser
from collections import defaultdict
import posixpath
from typing import Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

# XML 解析统一入口：安装了 defusedxml 时防御实体膨胀/外部实体等 XML 攻击，
# 未安装自动回退标准库 xml.etree（功能等价，仅防护降级）。
try:
    from defusedxml.ElementTree import fromstring as _safe_xml_fromstring
except ImportError:
    from xml.etree.ElementTree import fromstring as _safe_xml_fromstring


_FORBIDDEN_FILENAME_CHARS_RE = re.compile(r'[\\/:*?"<>|]')
_WHITESPACE_RE = re.compile(r"\s+")
_QUOTE_CHARS = '"“”‘’‚‛„‟«»‹›'


class DocxSecurityError(ValueError):
    """输入 DOCX 触发资源耗尽防线（zip bomb / 超大资源）。

    继承 ValueError 以兼容既有 except ValueError 处理；调用方可精确区分
    “安全拒绝”与普通格式/转换错误——安全拒绝表示输入恶意或异常，
    不可降级重试（若降级重试会绕过防线）。
    """


class ResourceLimitExceeded(DocxSecurityError):
    """可降级资源限制超限（图片数量/单图大小/单图像素/嵌入 Excel 大小）。

    继承 DocxSecurityError：默认（on_limit="reject"）处置与安全拒绝完全
    一致——整篇拒绝、不降级不重试；仅当调用方显式选择 on_limit="skip"
    时，才用于精确捕获并降级为跳过该资源。ZIP bomb 等恶意特征（总量/
    单 entry 解压量/压缩比）始终抛基类 DocxSecurityError，任何模式下
    都不降级。
    """


# 资源耗尽防线阈值。集中为 dict 便于测试注入与按需收紧。
DOCX_SECURITY_LIMITS = {
    "total_uncompressed": 500 * 1024 * 1024,    # ZIP 总解压上限
    "entry_uncompressed": 100 * 1024 * 1024,    # 单 entry 解压上限
    "entry_ratio": 100,                         # 单 entry 压缩比上限
    "total_ratio": 100,                         # 总压缩比上限
    "total_ratio_min_compressed": 1024 * 1024,  # 总压缩比仅对压缩后 >1MB 的包判定
    "image_count": 500,                         # word/media 图片数量上限
    "image_file_size": 20 * 1024 * 1024,        # 单图文件大小上限
    "image_pixels": 50_000_000,                 # 单图像素上限（解压炸弹检测）
    "embedded_excel_size": 50 * 1024 * 1024,    # 嵌入 Excel 大小上限
}

# 批处理跳过判定用的完成标记文件名（JSON，记录源文件 SHA-256）
SENTINEL_FILENAME = ".converted"

# on_limit="skip" 时超限资源在 Markdown 中的可见说明文案。
# mammoth 回调用 __SKIPPED_IMAGE_<reason>__ 作为 src 占位，转换后统一
# 替换为对应说明，保证跳过在输出中可见且不引用不存在的资源文件。
SKIPPED_IMAGE_NOTE = {
    "size": "单图超过大小上限",
    "pixels": "单图像素超过上限",
    "count": "图片数量超过上限",
}


def validate_on_limit(on_limit: str) -> None:
    """校验公开 API 共用的资源超限处置参数。"""
    if on_limit not in ("reject", "skip"):
        raise ValueError(f"on_limit 仅支持 'reject' 或 'skip': {on_limit!r}")


def _fmt_bytes(size: int) -> str:
    if size >= 1024 * 1024 * 1024:
        return f"{size / (1024 * 1024 * 1024):.1f}GB"
    if size >= 1024 * 1024:
        return f"{size / (1024 * 1024):.1f}MB"
    return f"{size / 1024:.1f}KB"


def validate_docx_zip_security(zip_ref: zipfile.ZipFile, on_limit: str = "reject") -> None:
    """解压前依据 ZIP 中央目录元数据做资源耗尽防线校验。

    只读 compress_size/file_size 等声明值，不解压条目内容；超限抛
    DocxSecurityError。实际读取条目时另由 read_zip_entry_bounded /
    _read_media_image 在真实解压路径上兜底，防元数据谎报。

    防线分两层，on_limit 只影响第二层：
      1. ZIP bomb 等恶意特征（总解压量/单 entry 解压量/压缩比）无条件抛
         DocxSecurityError，任何模式下都不降级；
      2. 可降级资源（图片数量/单图大小/嵌入 Excel 大小的声明值检查）抛
         ResourceLimitExceeded；on_limit="skip" 时不在此抛出，改由提取
         阶段按真实读取逐项跳过（见 extract_content_from_docx）。
    """
    validate_on_limit(on_limit)
    limits = DOCX_SECURITY_LIMITS
    skip_mode = on_limit == "skip"
    total_compressed = 0
    total_uncompressed = 0
    image_count = 0
    seen_names = set()

    for info in zip_ref.infolist():
        if info.filename in seen_names:
            raise DocxSecurityError(f"ZIP 包含重复条目名，解压语义不唯一: {info.filename}")
        seen_names.add(info.filename)
        if info.is_dir():
            continue
        name = info.filename
        compressed = info.compress_size
        uncompressed = info.file_size
        total_compressed += compressed
        total_uncompressed += uncompressed

        if uncompressed > limits["entry_uncompressed"]:
            raise DocxSecurityError(
                f"ZIP 条目解压后超过单文件上限 {_fmt_bytes(limits['entry_uncompressed'])}: "
                f"{name}（{_fmt_bytes(uncompressed)}）"
            )
        if compressed > 0 and uncompressed > compressed * limits["entry_ratio"]:
            raise DocxSecurityError(
                f"ZIP 条目压缩比超过 {limits['entry_ratio']}x: {name}"
            )

        if name.startswith("word/media/"):
            image_count += 1
            if not skip_mode and uncompressed > limits["image_file_size"]:
                raise ResourceLimitExceeded(
                    f"图片超过单图大小上限 {_fmt_bytes(limits['image_file_size'])}: {name}"
                )
        if name.startswith("word/embeddings/") and name.lower().endswith(".xlsx"):
            if not skip_mode and uncompressed > limits["embedded_excel_size"]:
                raise ResourceLimitExceeded(
                    f"嵌入 Excel 超过大小上限 {_fmt_bytes(limits['embedded_excel_size'])}: {name}"
                )

    if not skip_mode and image_count > limits["image_count"]:
        raise ResourceLimitExceeded(f"图片数量超过上限 {limits['image_count']}: 实际 {image_count} 张")
    if total_uncompressed > limits["total_uncompressed"]:
        raise DocxSecurityError(
            f"ZIP 总解压大小超过上限 {_fmt_bytes(limits['total_uncompressed'])}: "
            f"实际 {_fmt_bytes(total_uncompressed)}"
        )
    if total_compressed > limits["total_ratio_min_compressed"]:
        if total_uncompressed > total_compressed * limits["total_ratio"]:
            raise DocxSecurityError(
                f"ZIP 总压缩比超过 {limits['total_ratio']}x: "
                f"{_fmt_bytes(total_uncompressed)}/{_fmt_bytes(total_compressed)}"
            )


def read_zip_entry_bounded(
    zip_ref: zipfile.ZipFile, name: str, max_bytes: int,
    error_cls: type = DocxSecurityError,
) -> bytes:
    """带实际上限的条目读取：边解压边计数，超过 max_bytes 立即中止。

    validate_docx_zip_security 依赖 ZIP 声明值，本函数在真实解压路径上
    再兜一道底，防止中央目录元数据与实际数据不一致的恶意构造。
    error_cls 允许可降级资源（图片/嵌入 Excel）抛 ResourceLimitExceeded，
    供 on_limit="skip" 精确捕获降级；默认 DocxSecurityError 保持既有语义。
    """
    chunks = []
    total = 0
    with zip_ref.open(name) as fh:
        while True:
            chunk = fh.read(1024 * 1024)
            if not chunk:
                break
            total += len(chunk)
            if total > max_bytes:
                raise error_cls(
                    f"ZIP 条目实际解压超过上限 {_fmt_bytes(max_bytes)}: {name}"
                )
            chunks.append(chunk)
    return b"".join(chunks)


def _png_dimensions(data: bytes) -> Optional[Tuple[int, int]]:
    if len(data) < 24 or data[12:16] != b"IHDR":
        return None
    return int.from_bytes(data[16:20], "big"), int.from_bytes(data[20:24], "big")


_JPEG_SOF_MARKERS = frozenset(range(0xC0, 0xD0)) - {0xC4, 0xC8, 0xCC}


def _jpeg_dimensions(data: bytes) -> Optional[Tuple[int, int]]:
    i = 2
    length = len(data)
    while i + 4 <= length:
        if data[i] != 0xFF:
            i += 1
            continue
        marker = data[i + 1]
        if marker == 0xFF:
            i += 1
            continue
        if marker in (0xD8, 0xD9, 0xDA):  # SOI/EOI/SOS：SOF 应在 SOS 之前出现
            return None
        if 0xD0 <= marker <= 0xD7 or marker == 0x01:
            i += 2
            continue
        seg_len = int.from_bytes(data[i + 2:i + 4], "big")
        if seg_len < 2:
            return None
        if marker in _JPEG_SOF_MARKERS and i + 9 <= length:
            height = int.from_bytes(data[i + 5:i + 7], "big")
            width = int.from_bytes(data[i + 7:i + 9], "big")
            return width, height
        i += 2 + seg_len
    return None


def _gif_dimensions(data: bytes) -> Optional[Tuple[int, int]]:
    if len(data) < 10:
        return None
    return int.from_bytes(data[6:8], "little"), int.from_bytes(data[8:10], "little")


def _bmp_dimensions(data: bytes) -> Optional[Tuple[int, int]]:
    if len(data) < 26:
        return None
    width = int.from_bytes(data[18:22], "little", signed=True)
    height = abs(int.from_bytes(data[22:26], "little", signed=True))
    return width, height


def _webp_dimensions(data: bytes) -> Optional[Tuple[int, int]]:
    if len(data) < 30:
        return None
    chunk = data[12:16]
    if chunk == b"VP8X":
        # 扩展格式：画布宽高各 24bit，存储的是实际值-1
        return (int.from_bytes(data[24:27], "little") + 1,
                int.from_bytes(data[27:30], "little") + 1)
    if chunk == b"VP8 ":
        # 有损格式：keyframe 宽高各 14bit
        return (int.from_bytes(data[26:28], "little") & 0x3FFF,
                int.from_bytes(data[28:30], "little") & 0x3FFF)
    if chunk == b"VP8L" and len(data) >= 25:
        # 无损格式：签名字节后宽高各 14bit，存储的是实际值-1
        bits = int.from_bytes(data[21:25], "little")
        return (bits & 0x3FFF) + 1, ((bits >> 14) & 0x3FFF) + 1
    return None


def _tiff_dimensions(data: bytes) -> Optional[Tuple[int, int]]:
    if len(data) < 8:
        return None
    if data[:2] == b"II":
        endian = "little"
    elif data[:2] == b"MM":
        endian = "big"
    else:
        return None
    ifd_offset = int.from_bytes(data[4:8], endian)
    if ifd_offset + 2 > len(data):
        return None
    entry_count = int.from_bytes(data[ifd_offset:ifd_offset + 2], endian)
    width = height = None
    for idx in range(entry_count):
        base = ifd_offset + 2 + idx * 12
        if base + 12 > len(data):
            break
        tag = int.from_bytes(data[base:base + 2], endian)
        if tag not in (256, 257):  # ImageWidth / ImageLength
            continue
        vtype = int.from_bytes(data[base + 2:base + 4], endian)
        raw = data[base + 8:base + 12]
        if vtype == 3:  # SHORT
            value = int.from_bytes(raw[:2], endian)
        elif vtype == 4:  # LONG
            value = int.from_bytes(raw[:4], endian)
        elif vtype in (16, 17):  # LONG8 / SLONG8
            value = int.from_bytes(raw[:8], endian)
        else:
            continue
        if tag == 256:
            width = value
        else:
            height = value
    if width and height:
        return width, height
    return None


def image_pixel_count(image_data: bytes) -> Optional[int]:
    """从图片头部解析宽高并返回像素数（宽×高）；无法解析返回 None。

    用于解压炸弹（decompression bomb）检测：仅需头部几十字节即可判定，
    不解码像素数据。WMF/EMF 等矢量格式无固定位图像素，返回 None。
    """
    data = image_data or b""
    dims: Optional[Tuple[int, int]] = None
    if data[:8] == b'\x89PNG\r\n\x1a\n':
        dims = _png_dimensions(data)
    elif data[:2] == b'\xff\xd8':
        dims = _jpeg_dimensions(data)
    elif data[:6] in (b'GIF87a', b'GIF89a'):
        dims = _gif_dimensions(data)
    elif data[:2] == b'BM':
        dims = _bmp_dimensions(data)
    elif data[:4] == b'RIFF' and data[8:12] == b'WEBP':
        dims = _webp_dimensions(data)
    elif data[:4] in (b'II*\x00', b'MM\x00*'):
        dims = _tiff_dimensions(data)
    if not dims:
        return None
    width, height = dims
    if width <= 0 or height <= 0:
        return None
    return width * height


def _read_media_image(zip_ref: zipfile.ZipFile, name: str) -> bytes:
    """读取 word/media 图片并执行大小/像素防线（真实解压路径的兜底）。

    大小与像素均属可降级资源限制，抛 ResourceLimitExceeded：reject 模式
    下与 DocxSecurityError 处置一致（整篇拒绝），skip 模式下由调用方
    捕获并跳过该图片。
    """
    limits = DOCX_SECURITY_LIMITS
    image_data = read_zip_entry_bounded(
        zip_ref, name, limits["image_file_size"], error_cls=ResourceLimitExceeded
    )
    pixels = image_pixel_count(image_data)
    if pixels is not None and pixels > limits["image_pixels"]:
        raise ResourceLimitExceeded(
            f"图片像素超过上限 {limits['image_pixels']}: {name}（{pixels} 像素）"
        )
    return image_data


def sha256_file(path: str, chunk_size: int = 1024 * 1024) -> str:
    """流式计算文件 SHA-256，避免大文件一次性读入内存。"""
    digest = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(chunk_size), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_conversion_sentinel(
    final_output_dir: str, folder_name: str, source_sha256: str,
    on_limit: str = "reject",
) -> None:
    """原子写入转换完成标记（tmp + rename），记录输出目录名与源文件哈希。

    批处理据此判断“输出完整且与当前源一致”；仅转换全部成功后调用。
    标记写失败不影响本次转换结果，仅意味着批处理下次会重转。
    """
    sentinel_path = os.path.join(final_output_dir, SENTINEL_FILENAME)
    tmp_path = sentinel_path + ".tmp"
    validate_on_limit(on_limit)
    payload = json.dumps(
        {"folder_name": folder_name, "source_sha256": source_sha256, "on_limit": on_limit},
        ensure_ascii=False,
        sort_keys=True,
    )
    try:
        with open(tmp_path, "w", encoding="utf-8") as f:
            f.write(payload)
        os.replace(tmp_path, sentinel_path)
    except OSError:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        logger.warning("写入完成标记失败: %s", sentinel_path, exc_info=True)


def read_conversion_sentinel(directory: str) -> Optional[Dict[str, str]]:
    """读取完成标记；缺失/损坏/旧格式（纯文本等非 JSON 对象）返回 None。"""
    sentinel_path = os.path.join(directory, SENTINEL_FILENAME)
    try:
        with open(sentinel_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, ValueError):
        return None
    if not isinstance(data, dict):
        return None
    folder_name = data.get("folder_name")
    source_sha256 = data.get("source_sha256")
    # V0.1.6 sentinel 没有策略字段，当时唯一行为是 reject。
    on_limit = data.get("on_limit", "reject")
    if not isinstance(folder_name, str) or not folder_name:
        return None
    if not isinstance(source_sha256, str) or not source_sha256:
        return None
    if on_limit not in ("reject", "skip"):
        return None
    return {"folder_name": folder_name, "source_sha256": source_sha256, "on_limit": on_limit}


def prune_stale_assets(assets_dir: str, current_image_sources) -> None:
    """成功转换后删除当前结果不再使用的旧资源文件。

    单文件 API/CLI 允许复用同名输出目录。只在 Markdown 已成功生成后执行，
    避免失败转换提前破坏旧结果；目录及其子目录不删除。
    """
    current_names = {
        os.path.basename(source)
        for source in current_image_sources
        if isinstance(source, str) and source
    }
    with os.scandir(assets_dir) as entries:
        for entry in entries:
            if entry.name in current_names:
                continue
            if entry.is_symlink() or entry.is_file(follow_symlinks=False):
                os.remove(entry.path)


def _mammoth_embedded_media_key(image) -> Optional[str]:
    """尽力从 Mammoth 图片打开函数中取出嵌入媒体路径。

    Mammoth 的公开 Image 对象不暴露 relationship/path，但当前稳定
    实现会在 open 闭包中捕获 `word/media/...` 路径。只将明确识别
    的 DOCX 内嵌媒体用于物理条目去重；链接图片不归入 ZIP 配额。
    """
    opener = getattr(image, "open", None)
    closure = getattr(opener, "__closure__", None) or ()
    for cell in closure:
        try:
            value = cell.cell_contents
        except ValueError:
            continue
        if isinstance(value, str):
            normalized = posixpath.normpath(value.replace("\\", "/"))
            if normalized.startswith("word/media/"):
                return normalized
    return None


def _normalize_markdown_cell_text(value: str) -> str:
    """将单元格内容规范化为 Markdown 管道表可安全呈现的单行文本。"""
    text = unescape(value or "").replace("\xa0", " ")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.split("\n")]
    lines = [line for line in lines if line]
    text = "<br>".join(lines) if lines else ""
    return text.replace("|", r"\|")


class _TableHTMLParser(HTMLParser):
    """解析单个 HTML table，保留 rowspan/colspan 与单元格文本。"""

    def __init__(self):
        super().__init__()
        self.rows = []
        self._in_tr = False
        self._in_cell = False
        self._current_row = []
        self._cell_parts = []
        self._cell_tag = None
        self._cell_rowspan = 1
        self._cell_colspan = 1

    @staticmethod
    def _safe_int(raw, default=1):
        try:
            value = int(raw)
            return value if value > 0 else default
        except Exception:
            return default

    def handle_starttag(self, tag, attrs):
        attrs_map = dict(attrs)
        tag = tag.lower()
        if tag == "tr":
            self._in_tr = True
            self._current_row = []
            return

        if tag in ("td", "th") and self._in_tr:
            self._in_cell = True
            self._cell_tag = tag
            self._cell_parts = []
            self._cell_rowspan = self._safe_int(attrs_map.get("rowspan"), 1)
            self._cell_colspan = self._safe_int(attrs_map.get("colspan"), 1)
            return

        if self._in_cell and tag in ("br",):
            self._cell_parts.append("\n")
        elif self._in_cell and tag in ("p", "div", "li"):
            if self._cell_parts and not self._cell_parts[-1].endswith("\n"):
                self._cell_parts.append("\n")

    def handle_endtag(self, tag):
        tag = tag.lower()
        if tag in ("p", "div", "li") and self._in_cell:
            if self._cell_parts and not self._cell_parts[-1].endswith("\n"):
                self._cell_parts.append("\n")
            return

        if tag in ("td", "th") and self._in_cell:
            text = _normalize_markdown_cell_text("".join(self._cell_parts))
            self._current_row.append(
                {
                    "text": text,
                    "rowspan": self._cell_rowspan,
                    "colspan": self._cell_colspan,
                    "is_header": self._cell_tag == "th",
                }
            )
            self._in_cell = False
            self._cell_parts = []
            self._cell_tag = None
            self._cell_rowspan = 1
            self._cell_colspan = 1
            return

        if tag == "tr" and self._in_tr:
            if self._current_row:
                self.rows.append(self._current_row)
            self._in_tr = False
            self._current_row = []

    def handle_data(self, data):
        if self._in_cell:
            self._cell_parts.append(data)


def _normalize_list_item_text(value: str) -> str:
    lines = []
    for raw in (value or "").replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        if not raw.strip():
            continue
        # 嵌套列表行保留原始缩进，避免层级被破坏。
        if re.match(r"^\s+(-|\d+\.)\s+", raw):
            lines.append(raw.rstrip())
        else:
            lines.append(re.sub(r"\s+", " ", raw).strip())
    return "\n".join(lines)


class _ListHTMLTransformer(HTMLParser):
    """将 HTML 列表结构转换为 Markdown 列表，保留嵌套层级。"""

    def __init__(self):
        super().__init__()
        self._out = []
        self._list_stack = []  # [{"type": "ul"/"ol", "items": [str, ...]}]
        self._li_stack = []  # [list[str], ...]

    @staticmethod
    def _attrs_to_str(attrs):
        if not attrs:
            return ""
        pairs = []
        for k, v in attrs:
            if v is None:
                pairs.append(k)
            else:
                escaped = str(v).replace('"', "&quot;")
                pairs.append(f'{k}="{escaped}"')
        return " " + " ".join(pairs)

    @staticmethod
    def _render_list(context, depth):
        indent = "  " * depth
        is_ordered = context["type"] == "ol"
        lines = []
        for idx, item in enumerate(context["items"], 1):
            marker = f"{idx}. " if is_ordered else "- "
            normalized = _normalize_list_item_text(item)
            if not normalized:
                lines.append(f"{indent}{marker}".rstrip())
                continue
            item_lines = normalized.splitlines()
            lines.append(f"{indent}{marker}{item_lines[0].strip()}")
            for extra in item_lines[1:]:
                if extra.startswith("  "):
                    lines.append(f"{indent}{extra}")
                else:
                    lines.append(f"{indent}  {extra.strip()}")
        return "\n".join(lines)

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()
        if tag in ("ul", "ol"):
            self._list_stack.append({"type": tag, "items": []})
            return
        if tag == "li" and self._list_stack:
            self._li_stack.append([])
            return

        if self._li_stack:
            if tag == "br":
                self._li_stack[-1].append("\n")
            elif tag in ("p", "div"):
                if self._li_stack[-1] and not self._li_stack[-1][-1].endswith("\n"):
                    self._li_stack[-1].append("\n")
            return

        self._out.append(f"<{tag}{self._attrs_to_str(attrs)}>")

    def handle_endtag(self, tag):
        tag = tag.lower()
        if tag in ("ul", "ol") and self._list_stack:
            context = self._list_stack.pop()
            md = self._render_list(context, len(self._list_stack))
            if self._li_stack:
                if self._li_stack[-1] and not self._li_stack[-1][-1].endswith("\n"):
                    self._li_stack[-1].append("\n")
                self._li_stack[-1].append(md)
            else:
                # 顶层列表后补空行，避免后续表格被当作列表延续文本。
                self._out.append("\n" + md + "\n\n")
            return

        if tag == "li" and self._li_stack:
            item_text = "".join(self._li_stack.pop())
            if self._list_stack:
                self._list_stack[-1]["items"].append(item_text)
            return

        if self._li_stack:
            if tag in ("p", "div"):
                if self._li_stack[-1] and not self._li_stack[-1][-1].endswith("\n"):
                    self._li_stack[-1].append("\n")
            return

        self._out.append(f"</{tag}>")

    def handle_startendtag(self, tag, attrs):
        tag = tag.lower()
        if self._li_stack and tag == "br":
            self._li_stack[-1].append("\n")
            return
        if self._li_stack:
            return
        self._out.append(f"<{tag}{self._attrs_to_str(attrs)}/>")

    def handle_data(self, data):
        if self._li_stack:
            self._li_stack[-1].append(data)
            return
        if self._list_stack:
            # 列表容器中但不在 li 内的噪声文本通常只有空白，忽略。
            return
        self._out.append(data)

    def get_output(self):
        return "".join(self._out)


def transform_html_lists_to_markdown(html: str) -> str:
    parser = _ListHTMLTransformer()
    parser.feed(html)
    parser.close()
    return parser.get_output()


def _expand_table_rows(rows):
    """将包含 rowspan/colspan 的行展开为等宽二维表。"""
    expanded = []
    spans = {}  # col_idx -> {"rows_left": int, "text": str}

    for row in rows:
        out_row = []
        col = 0

        def consume_span_at_current_col():
            nonlocal col
            while col in spans:
                span = spans[col]
                out_row.append(span["text"])
                span["rows_left"] -= 1
                if span["rows_left"] <= 0:
                    spans.pop(col, None)
                col += 1

        consume_span_at_current_col()
        for cell in row:
            consume_span_at_current_col()
            text = cell["text"]
            rowspan = max(1, int(cell["rowspan"]))
            colspan = max(1, int(cell["colspan"]))
            for offset in range(colspan):
                out_row.append(text)
                if rowspan > 1:
                    spans[col + offset] = {"rows_left": rowspan - 1, "text": text}
            col += colspan

        consume_span_at_current_col()
        expanded.append(out_row)

    while spans:
        out_row = []
        col = 0
        max_col = max(spans.keys())
        while col <= max_col:
            if col in spans:
                span = spans[col]
                out_row.append(span["text"])
                span["rows_left"] -= 1
                if span["rows_left"] <= 0:
                    spans.pop(col, None)
            else:
                out_row.append("")
            col += 1
        expanded.append(out_row)

    width = max((len(row) for row in expanded), default=0)
    if width:
        expanded = [row + [""] * (width - len(row)) for row in expanded]
    return expanded


def table_html_to_markdown(table_html: str) -> str:
    parser = _TableHTMLParser()
    parser.feed(table_html)
    parser.close()

    rows = _expand_table_rows(parser.rows)
    if not rows:
        return ""

    lines = []
    lines.append("| " + " | ".join(rows[0]) + " |")
    lines.append("| " + " | ".join(["---"] * len(rows[0])) + " |")
    for row in rows[1:]:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines) + "\n\n"


def promote_numbered_bold_headings(markdown: str) -> str:
    """将“编号 + 加粗标题”段落提升为 Markdown 标题。"""
    pattern = re.compile(
        r"^(?P<num>\d+(?:\.\d+)*)(?P<dot>\.)?\s+\*\*(?P<title>[^*\n]+)\*\*\s*$",
        flags=0,
    )
    heading_pattern = re.compile(r"^(#{1,6})\s+")

    lines = markdown.splitlines()
    out = []
    previous_heading_level = 0
    previous_promoted_depth = None

    for line in lines:
        heading_match = heading_pattern.match(line)
        if heading_match:
            previous_heading_level = len(heading_match.group(1))
            previous_promoted_depth = None
            out.append(line)
            continue

        match = pattern.match(line.strip())
        if not match:
            out.append(line)
            continue

        num = match.group("num")
        dot = match.group("dot") or ""
        title = match.group("title").strip()
        depth = num.count(".") + 1
        level = min(depth, 6)  # 1级编号 -> #，2级编号 -> ##

        # 在深层章节下的“1. **小节**”更接近子标题，避免被抬到过高层级。
        if depth == 1 and previous_heading_level >= 2:
            if previous_promoted_depth == 1:
                level = previous_heading_level
            else:
                level = min(previous_heading_level + 1, 6)

        promoted = f"{'#' * level} {num}{dot} {title}"
        out.append(promoted)
        previous_heading_level = level
        previous_promoted_depth = depth

    # 保持原始编号，不做自动重排，避免双语并列标题或手工编号被误改。
    return "\n".join(out)


def promote_leading_bold_title(markdown: str) -> str:
    """将文档开头“整行加粗标题”提升为一级标题（保守触发）。"""
    lines = markdown.splitlines()
    first_idx = None
    for i, line in enumerate(lines):
        if line.strip():
            first_idx = i
            break
    if first_idx is None:
        return markdown

    first_line = lines[first_idx].strip()
    m = re.match(r"^\*\*(?P<title>.+?)\*\*$", first_line)
    if not m:
        return markdown

    # 仅在后续存在“编号章节标题”时触发，降低把普通强调段误判成标题的风险。
    section_heading_re = re.compile(r"^#{1,6}\s+\d+(?:\.\d+)*\.?\s+")
    has_numbered_section_heading = any(section_heading_re.match(line.strip()) for line in lines[first_idx + 1 :])
    if not has_numbered_section_heading:
        return markdown

    title = m.group("title").strip()
    if not title:
        return markdown

    lines[first_idx] = f"# {title}"
    return "\n".join(lines)


def sanitize_stem(stem: str) -> str:
    raw = stem  # 保留原始值用于 hash
    normalized = unicodedata.normalize("NFKC", raw or "")
    # NFKC 归一化（如全角→半角）不加 hash：中文文档场景过于普遍，
    # 由此产生的罕见碰撞由 sentinel 的源哈希校验兜底（不一致即重转）。
    no_quotes = normalized
    for ch in _QUOTE_CHARS:
        no_quotes = no_quotes.replace(ch, "")
    substituted = _FORBIDDEN_FILENAME_CHARS_RE.sub("_", no_quotes)
    # 引号删除或非法字符替换是强丢失映射（如 a:b 与 a_b 同映射），
    # 需附加 hash 防止不同原始名称共享同一输出目录
    lossy = no_quotes != normalized or substituted != no_quotes
    stem = _WHITESPACE_RE.sub(" ", substituted).strip()
    stem = stem.strip(". ").strip()
    if not stem:
        return "document"
    if len(stem) <= 120 and not lossy:
        return stem
    # 清洗发生丢失或超长截断时，附加原始全名的短 hash，避免不同文件名映射到同一输出目录
    suffix = hashlib.sha256(raw.encode("utf-8")).hexdigest()[:8]
    return f"{stem[:111]}_{suffix}"


def extract_heading_level_map(docx_path: str) -> Dict[str, int]:
    """解析 DOCX 的 heading bookmark 段落样式，映射为 Markdown 标题层级。"""
    ns_w = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    tag_p = f"{{{ns_w}}}p"
    tag_ppr = f"{{{ns_w}}}pPr"
    tag_pstyle = f"{{{ns_w}}}pStyle"
    tag_bm = f"{{{ns_w}}}bookmarkStart"
    attr_name = f"{{{ns_w}}}name"
    attr_val = f"{{{ns_w}}}val"
    tag_t = f"{{{ns_w}}}t"

    def style_to_level(style_val: str) -> Optional[int]:
        if not style_val:
            return None
        raw = str(style_val).strip()
        m = re.search(r"(\d+)$", raw)
        if not m:
            return None
        n = int(m.group(1))
        if n <= 0:
            return None
        # style=1/2/3 对应一/二/三级标题。
        return min(n, 6)

    def infer_level_from_text(text: str) -> Optional[int]:
        m = re.match(r"^\s*(\d+(?:\.\d+)*)\s*\.?\s+", text or "")
        if not m:
            return None
        depth = m.group(1).count(".") + 1
        return min(depth, 6)

    level_map: Dict[str, int] = {}
    try:
        with zipfile.ZipFile(docx_path, "r") as zip_ref:
            doc_xml = _safe_xml_fromstring(zip_ref.read("word/document.xml"))
        for p in doc_xml.findall(f".//{tag_p}"):
            bm = p.find(f".//{tag_bm}")
            if bm is None:
                continue
            name = bm.get(attr_name)
            if not name or not name.startswith("heading_"):
                continue

            ppr = p.find(tag_ppr)
            style_val = ""
            if ppr is not None:
                pstyle = ppr.find(tag_pstyle)
                if pstyle is not None:
                    style_val = pstyle.get(attr_val, "")

            level = style_to_level(style_val)
            if level is None:
                text = "".join((t.text or "") for t in p.findall(f".//{tag_t}"))
                level = infer_level_from_text(text)
            if level is not None:
                level_map[name] = level
    except Exception:
        return {}

    return level_map


def resolve_part_path(target: str) -> str:
    """将 relationship target 解析为 docx zip 内的规范路径（如 word/media/image1.png）"""
    target = (target or "").replace("\\", "/").strip()
    if not target:
        return ""
    if target.startswith("/"):
        target = target[1:]
    if target.startswith("word/"):
        return posixpath.normpath(target)
    return posixpath.normpath(posixpath.join("word", target))


def parse_relationships(docx_path):
    """解析docx中的关系文件，找出Excel嵌入和对应预览图的映射。

    策略：
      1. 优先从 document.xml 中解析 <w:object> 节点，提取 OLEObject rId
         和 imagedata rId 的真实配对关系（最可靠）。
      2. 对方法1未覆盖的项，使用 "rId相邻" 启发式补全（兼容）。
    """
    excel_to_preview = {}  # Excel路径 -> 预览图路径
    preview_to_excel = {}  # 预览图路径 -> Excel路径
    ordered_pairs = []  # [(Excel路径, 预览图路径)]，按文档出现顺序

    # --- 公共：解析 rels 文件，建立 rId -> target 映射 ---
    NS_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
    relationships = {}  # rId -> {'type': ..., 'target': ...}

    with zipfile.ZipFile(docx_path, 'r') as zip_ref:
        try:
            rels_content = zip_ref.read('word/_rels/document.xml.rels')
            rels_root = _safe_xml_fromstring(rels_content)
            for rel in rels_root.findall(f'.//{{{NS_REL}}}Relationship'):
                rid = rel.get('Id')
                rel_type = rel.get('Type', '').split('/')[-1]
                target = rel.get('Target', '')
                relationships[rid] = {'type': rel_type, 'target': target}
        except Exception as e:
            logger.warning("解析关系文件失败: %s", e)
            return excel_to_preview, preview_to_excel, ordered_pairs

        # --- 方法1：从 document.xml 解析 OLE 对象的真实引用 ---
        NS_W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
        NS_V = "urn:schemas-microsoft-com:vml"
        NS_O = "urn:schemas-microsoft-com:office:office"
        NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

        try:
            doc_xml = zip_ref.read('word/document.xml')
            doc_root = _safe_xml_fromstring(doc_xml)

            # 查找所有 <w:object> 节点（可能嵌套在 mc:AlternateContent 等下面）
            for obj_node in doc_root.iter(f'{{{NS_W}}}object'):
                ole_rid = None
                img_rid = None

                # <o:OLEObject r:id="rIdX" />
                for ole in obj_node.iter(f'{{{NS_O}}}OLEObject'):
                    ole_rid = ole.get(f'{{{NS_R}}}id')

                # <v:imagedata r:id="rIdY" />
                for imgdata in obj_node.iter(f'{{{NS_V}}}imagedata'):
                    img_rid = imgdata.get(f'{{{NS_R}}}id')

                if ole_rid and img_rid and ole_rid in relationships and img_rid in relationships:
                    ole_target = resolve_part_path(relationships[ole_rid]['target'])
                    img_target = resolve_part_path(relationships[img_rid]['target'])
                    if ole_target.lower().endswith('.xlsx'):
                        excel_to_preview[ole_target] = img_target
                        preview_to_excel[img_target] = ole_target
                        ordered_pairs.append((ole_target, img_target))
        except Exception:
            pass  # document.xml 解析失败不影响后续

        # --- 方法2（补全）：rId 相邻启发式，补全方法1未覆盖的 Excel ---
        def rid_sort_key(rid: str) -> int:
            m = re.fullmatch(r"rId(\d+)", rid or "")
            return int(m.group(1)) if m else 10**9

        sorted_rids = sorted(relationships.keys(), key=rid_sort_key)

        for i, rid in enumerate(sorted_rids):
            rel = relationships[rid]
            if rel['type'] == 'package' and rel['target'].lower().endswith('.xlsx'):
                excel_file = resolve_part_path(rel['target'])
                if excel_file in excel_to_preview:
                    continue  # 已被方法1覆盖，跳过
                if i + 1 < len(sorted_rids):
                    next_rid = sorted_rids[i + 1]
                    next_rel = relationships[next_rid]
                    if next_rel['type'] == 'image':
                        preview_file = resolve_part_path(next_rel['target'])
                        excel_to_preview[excel_file] = preview_file
                        preview_to_excel[preview_file] = excel_file
                        ordered_pairs.append((excel_file, preview_file))

    return excel_to_preview, preview_to_excel, ordered_pairs


def _format_cell_value(cell) -> str:
    """将 openpyxl 单元格值转换为友好的字符串表示。"""
    if cell is None:
        return ''
    import datetime as _dt
    if isinstance(cell, _dt.datetime):
        if cell.hour == 0 and cell.minute == 0 and cell.second == 0 and cell.microsecond == 0:
            return cell.strftime("%Y-%m-%d")
        return cell.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(cell, _dt.date):
        return cell.strftime("%Y-%m-%d")
    if isinstance(cell, _dt.time):
        return cell.strftime("%H:%M:%S")
    if isinstance(cell, float) and math.isfinite(cell) and cell.is_integer():
        return str(int(cell))
    return str(cell)


def excel_to_markdown(xlsx_data):
    """将Excel数据转换为Markdown表格（仅依赖 openpyxl，无需 pandas）"""
    try:
        import openpyxl

        # XLSX 本身也是 ZIP：DOCX 外层的 entry 限制只能约束
        # xlsx 字节大小，无法防止其内部条目解压膨胀。交给
        # openpyxl 前先对内层 ZIP 无条件执行恶意特征校验。
        with zipfile.ZipFile(io.BytesIO(xlsx_data), "r") as xlsx_zip:
            validate_docx_zip_security(xlsx_zip, on_limit="reject")

        def normalize_rows(raw_rows: List[List[str]]) -> List[List[str]]:
            if not raw_rows:
                return []

            rows = [r for r in raw_rows if any(c.strip() for c in r)]
            if not rows:
                return []

            col_count = max(len(r) for r in rows)
            rows = [r + [''] * (col_count - len(r)) for r in rows]
            non_empty_cols = [j for j in range(col_count) if any(rows[i][j].strip() for i in range(len(rows)))]
            if not non_empty_cols:
                return []
            return [[r[j] for j in non_empty_cols] for r in rows]

        def apply_merged_cells(ws, raw_rows: List[List[str]]) -> List[List[str]]:
            """将合并单元格展开为 Markdown 管道表可读的全展开网格。"""
            if not raw_rows:
                return raw_rows

            for merged in ws.merged_cells.ranges:
                min_row, max_row = merged.min_row, merged.max_row
                min_col, max_col = merged.min_col, merged.max_col

                row_idx = min_row - 1
                col_idx = min_col - 1
                if row_idx >= len(raw_rows):
                    continue
                if col_idx >= len(raw_rows[row_idx]):
                    continue

                anchor_value = raw_rows[row_idx][col_idx]
                if not anchor_value:
                    continue

                for row_no in range(min_row, max_row + 1):
                    i = row_no - 1
                    if i >= len(raw_rows):
                        continue
                    if len(raw_rows[i]) < max_col:
                        raw_rows[i].extend([''] * (max_col - len(raw_rows[i])))
                    for col_no in range(min_col, max_col + 1):
                        raw_rows[i][col_no - 1] = anchor_value
            return raw_rows

        def sheet_to_rows(ws) -> tuple[List[List[str]], List[str], int]:
            raw_rows: List[List[str]] = []
            for row in ws.iter_rows(values_only=True):
                raw_rows.append([
                    _normalize_markdown_cell_text(_format_cell_value(cell))
                    for cell in row
                ])
            score_rows = normalize_rows(raw_rows)
            score = sum(1 for r in score_rows for c in r if c.strip())
            merge_ranges = [str(rng) for rng in ws.merged_cells.ranges]
            raw_rows = apply_merged_cells(ws, raw_rows)
            return normalize_rows(raw_rows), merge_ranges, score

        # 不使用 read_only=True：部分嵌入工作簿的维度元数据异常，read_only 模式会把表格截断成 1x1。
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_data), read_only=False, data_only=True)
        best_rows = []
        best_merge_ranges: List[str] = []
        best_score = -1
        for ws in wb.worksheets:
            rows, merge_ranges, score = sheet_to_rows(ws)
            if not rows:
                continue
            if score > best_score:
                best_rows = rows
                best_merge_ranges = merge_ranges
                best_score = score
        wb.close()

        if not best_rows:
            return None

        header = '| ' + ' | '.join(best_rows[0]) + ' |'
        separator = '| ' + ' | '.join(['---'] * len(best_rows[0])) + ' |'
        body_lines = ['| ' + ' | '.join(r) + ' |' for r in best_rows[1:]]
        table_text = header + '\n' + separator + '\n' + '\n'.join(body_lines)
        if best_merge_ranges:
            ranges_text = ", ".join(best_merge_ranges)
            return f"> merge_ranges: {ranges_text}\n\n{table_text}"
        return table_text

    except DocxSecurityError:
        raise
    except Exception as e:
        logger.warning("Excel转Markdown失败: %s", e)
        return None


def detect_image_format(image_data):
    """检测图片的真实格式；无法识别时返回 None（调用方应保留原扩展名）"""
    if image_data[:8] == b'\x89PNG\r\n\x1a\n':
        return '.png'
    elif image_data[:2] == b'\xff\xd8':
        return '.jpeg'
    elif image_data[:6] in (b'GIF87a', b'GIF89a'):
        return '.gif'
    elif image_data[:4] == b'RIFF' and image_data[8:12] == b'WEBP':
        return '.webp'
    elif image_data[:2] == b'BM':
        return '.bmp'
    elif image_data[:4] in (b'II*\x00', b'MM\x00*'):
        return '.tiff'
    elif image_data[:4] == b'\xd7\xcd\xc6\x9a':
        return '.wmf'
    elif len(image_data) >= 44 and image_data[40:44] == b' EMF':
        return '.emf'
    return None


def extract_content_from_docx(docx_path, assets_dir, on_limit="reject", skip_state=None):
    """从docx中提取图片和Excel数据，并构建“内容hash -> 内容”的映射

    Args:
        on_limit: 可降级资源（图片数量/单图大小/单图像素/嵌入 Excel 大小）
            超限时的处置。"reject"（默认）抛 ResourceLimitExceeded 整篇拒绝；
            "skip" 仅跳过该资源继续转换。
        skip_state: 可选的可变 dict 输出参数。传入时写入 skipped 与
            media_processed（配额内的媒体条目数）、allowed_media_paths
            （提取与 Mammoth 回调共用的物理媒体白名单）；省略时保持
            历史三项返回值契约。
            ZIP 级恶意特征不在此降级（validate 阶段已无条件拒绝）。

    返回:
        image_by_hash: { sha256_hex: "assets/xxx.png" }
        table_queue_by_hash: { sha256_hex: ["<md_table1>", "<md_table2>", ...] }
        table_repeat_by_hash: { sha256_hex: "<md_table>" }  # 队列耗尽时的稳定兜底
    """
    validate_on_limit(on_limit)
    skip_mode = on_limit == "skip"
    limits = DOCX_SECURITY_LIMITS
    skipped = []
    media_processed = 0
    image_by_hash = {}
    table_queue_by_hash = defaultdict(list)
    table_repeat_by_hash = {}

    # 解析关系，找出Excel和预览图的对应
    excel_to_preview, preview_to_excel, ordered_pairs = parse_relationships(docx_path)

    with zipfile.ZipFile(docx_path, 'r') as zip_ref:
        excel_md_by_path = {}
        table_preview_paths = set()
        media_paths = [
            info.filename for info in zip_ref.filelist
            if info.filename.startswith('word/media/') and not info.is_dir()
        ]
        allowed_media_paths = set(media_paths)
        if skip_mode:
            allowed_media_paths = set(media_paths[:limits["image_count"]])
            media_processed = len(allowed_media_paths)
            if len(media_paths) > limits["image_count"]:
                skipped.append((
                    "word/media/*",
                    f"图片数量超过上限 {limits['image_count']}，剩余图片停止提取",
                ))
                logger.warning(
                    "图片数量超过上限 %d，配额外 word/media 图片不会读取",
                    limits["image_count"],
                )

        # 先提取所有 Excel 文件的数据并转换为 Markdown
        for file_info in zip_ref.filelist:
            if file_info.filename.startswith('word/embeddings/') and file_info.filename.lower().endswith('.xlsx'):
                excel_file = file_info.filename
                try:
                    xlsx_data = read_zip_entry_bounded(
                        zip_ref, excel_file, limits["embedded_excel_size"],
                        error_cls=ResourceLimitExceeded,
                    )
                except ResourceLimitExceeded:
                    if not skip_mode:
                        raise
                    skipped.append((excel_file, "嵌入 Excel 超过大小上限"))
                    logger.warning("跳过超大嵌入 Excel（表格不转换，正文保留）: %s", excel_file)
                    continue

                markdown_table = excel_to_markdown(xlsx_data)
                if markdown_table:
                    excel_md_by_path[excel_file] = markdown_table
                else:
                    logger.warning("Excel表格转换失败（将保留预览图）: %s", excel_file)

        # 建立预览图 hash -> 表格队列（同一预览图内容可对应多个表格）
        pairs = ordered_pairs if ordered_pairs else [(e, p) for e, p in excel_to_preview.items()]
        for excel_path, preview_path in pairs:
            table_md = excel_md_by_path.get(excel_path)
            if not table_md:
                continue
            if preview_path not in zip_ref.namelist():
                continue
            table_preview_paths.add(preview_path)
            if skip_mode and preview_path not in allowed_media_paths:
                continue
            try:
                preview_data = _read_media_image(zip_ref, preview_path)
            except ResourceLimitExceeded as exc:
                if not skip_mode:
                    raise
                skipped.append((preview_path, str(exc)))
                logger.warning("跳过超限预览图（对应表格不注入，原位置显示跳过说明）: %s", exc)
                continue
            digest = hashlib.sha256(preview_data).hexdigest()
            table_queue_by_hash[digest].append(table_md)
            table_repeat_by_hash[digest] = table_md
            logger.info("转换Excel为表格: %s", excel_path)

        # 处理图片（显式目录 entry 不算图片，否则会写出空 assets/.png 并错占配额）
        for file_info in zip_ref.filelist:
            if file_info.filename.startswith('word/media/') and not file_info.is_dir():
                image_name = os.path.basename(file_info.filename)

                # 检查这个图片是否是Excel的预览图
                if file_info.filename in table_preview_paths:
                    continue

                if skip_mode and file_info.filename not in allowed_media_paths:
                    continue

                # 普通图片，直接提取（读取路径上执行大小/像素防线）
                try:
                    image_data = _read_media_image(zip_ref, file_info.filename)
                except ResourceLimitExceeded as exc:
                    if not skip_mode:
                        raise
                    skipped.append((file_info.filename, str(exc)))
                    logger.warning("跳过超限图片: %s", exc)
                    continue
                digest = hashlib.sha256(image_data).hexdigest()
                
                # 检测真实的图片格式并修正扩展名；无法识别时保留原扩展名，
                # 避免把 WMF/EMF 等格式误写成 .png 造成文件损坏
                original_ext = os.path.splitext(image_name)[1].lower()
                actual_ext = detect_image_format(image_data) or original_ext or ".png"
                base_name = os.path.splitext(image_name)[0]
                corrected_name = f"{base_name}{actual_ext}"
                
                image_path = os.path.join(assets_dir, corrected_name)
                # 扩展名修正后可能与已有文件同名，若内容不同则附加hash后缀避免覆盖
                if os.path.exists(image_path):
                    try:
                        with open(image_path, "rb") as f:
                            existing = f.read()
                        if existing != image_data:
                            corrected_name = f"{base_name}_{digest[:8]}{actual_ext}"
                            image_path = os.path.join(assets_dir, corrected_name)
                    except Exception:
                        corrected_name = f"{base_name}_{digest[:8]}{actual_ext}"
                        image_path = os.path.join(assets_dir, corrected_name)

                if not os.path.exists(image_path):
                    with open(image_path, 'wb') as f:
                        f.write(image_data)

                image_by_hash.setdefault(digest, f"assets/{corrected_name}")
                logger.info("提取图片: %s", corrected_name)
    
    if skip_state is not None:
        skip_state.clear()
        skip_state.update({
            "skipped": skipped,
            "media_processed": media_processed,
            "allowed_media_paths": allowed_media_paths,
        })
    return image_by_hash, table_queue_by_hash, table_repeat_by_hash


def extract_textbox_content(docx_path: str) -> List[str]:
    """从 DOCX 的 document.xml 中提取文本框 (<w:txbxContent>) 内的纯文本。

    mammoth 通常会忽略 text box / shape 中的内容，此函数作为补充。
    返回非空文本块列表。
    """
    ns_w = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    ns_wps = "http://schemas.microsoft.com/office/word/2010/wordprocessingShape"
    tag_txbx = f"{{{ns_w}}}txbxContent"
    tag_txbx_wps = f"{{{ns_wps}}}txbxContent"
    tag_t = f"{{{ns_w}}}t"
    tag_p = f"{{{ns_w}}}p"

    blocks: List[str] = []
    try:
        with zipfile.ZipFile(docx_path, "r") as zf:
            doc_xml = _safe_xml_fromstring(zf.read("word/document.xml"))

        for txbx_tag in (tag_txbx, tag_txbx_wps):
            for txbx in doc_xml.iter(txbx_tag):
                paras = []
                for p in txbx.findall(f".//{tag_p}"):
                    text = "".join((t.text or "") for t in p.findall(f".//{tag_t}"))
                    text = text.strip()
                    if text:
                        paras.append(text)
                if paras:
                    blocks.append("\n".join(paras))
    except Exception:
        pass
    return blocks


def extract_math_text(docx_path: str) -> List[str]:
    """从 DOCX 的 document.xml 中提取 OMML 数学公式的纯文本内容。

    完整的 OMML→LaTeX 转换极为复杂，此函数仅提取公式中的文本节点，
    用 $ 包裹作为占位标记，便于下游人工校正。
    """
    ns_m = "http://schemas.openxmlformats.org/officeDocument/2006/math"
    ns_w = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    tag_omath = f"{{{ns_m}}}oMath"
    tag_omath_para = f"{{{ns_m}}}oMathPara"
    tag_t_m = f"{{{ns_m}}}t"
    tag_t_w = f"{{{ns_w}}}t"

    formulas: List[str] = []
    try:
        with zipfile.ZipFile(docx_path, "r") as zf:
            doc_xml = _safe_xml_fromstring(zf.read("word/document.xml"))

        seen = set()
        for parent_tag in (tag_omath_para, tag_omath):
            for node in doc_xml.iter(parent_tag):
                node_id = id(node)
                if node_id in seen:
                    continue
                seen.add(node_id)
                parts = []
                for t in node.iter():
                    if t.tag in (tag_t_m, tag_t_w) and t.text:
                        parts.append(t.text)
                text = "".join(parts).strip()
                if text:
                    formulas.append(text)
                for child in node.iter(tag_omath):
                    seen.add(id(child))
    except Exception:
        pass
    return formulas


def convert_docx_to_markdown(docx_path, output_dir, create_subfolder=True, output_name=None,
                             on_limit="reject"):
    """将docx转换为markdown

    Args:
        docx_path: DOCX 文件路径
        output_dir: 输出目录路径
        create_subfolder: 是否在输出目录下创建以文件名命名的子文件夹（默认 True）
        output_name: 自定义输出命名（默认 None 用源文件名）。末尾 .docx 自动去除，
            其他点号后缀保留。
            经 sanitize_stem 清洗后统一用于子文件夹名、.md 文件名与 sentinel 的
            folder_name 字段，三处保持一致。适合 Web 上传等需要以用户原始
            文件名命名的场景；批处理不使用本参数（按源文件名命名）
        on_limit: 可降级资源（图片数量/单图大小/单图像素/嵌入 Excel 大小）超限
            处置。"reject"（默认）抛 DocxSecurityError 整篇拒绝，与历史行为一致；
            "skip" 仅跳过超限资源继续转换（超限图片原位置写入可见跳过说明，
            不落盘）。ZIP bomb 等恶意特征在任何模式下都整篇拒绝
    """
    validate_on_limit(on_limit)
    skip_mode = on_limit == "skip"
    
    # 先校验输入，避免 BadZipFile 直接中断并泄漏底层异常。
    # 安全校验（DocxSecurityError）在结构校验之后执行：结构非法报格式错误，
    # 结构合法但资源超限报安全错误。二者均为 ValueError 子类，上层可按需
    # 区分——安全错误表示输入恶意/异常，不可降级重试。
    try:
        with zipfile.ZipFile(docx_path, "r") as zip_ref:
            if "word/document.xml" not in zip_ref.namelist():
                raise ValueError(f"输入文件不是有效的 DOCX（缺少 word/document.xml）: {docx_path}")
            validate_docx_zip_security(zip_ref, on_limit=on_limit)
    except zipfile.BadZipFile as exc:
        raise ValueError(f"输入文件不是有效的 DOCX/ZIP: {docx_path}") from exc

    # 记录源文件哈希：转换全部成功后写入 .converted sentinel，
    # 供批处理判断“输出完整且与当前源一致”（源变更后自动重转）。
    source_sha256 = sha256_file(docx_path)

    # 输出命名：优先显式指定的 output_name（如 Web 上传场景的用户原始文件名），
    # 统一经 sanitize_stem 清洗后作为 folder_name，下游目录/文件名/sentinel 均引用它
    if output_name is not None:
        if not isinstance(output_name, str) or not output_name.strip():
            raise ValueError("output_name 必须是非空字符串")
        normalized_output_name = output_name.strip()
        if normalized_output_name.lower().endswith(".docx"):
            normalized_output_name = normalized_output_name[:-5]
        if not normalized_output_name.strip():
            raise ValueError("output_name 去除 .docx 后不能为空")
        folder_name = sanitize_stem(normalized_output_name)
    else:
        base_name = os.path.splitext(os.path.basename(docx_path))[0]
        folder_name = sanitize_stem(base_name)
    
    # 确定最终输出目录
    if create_subfolder:
        final_output_dir = os.path.join(output_dir, folder_name)
    else:
        final_output_dir = output_dir
    
    # 创建输出目录
    os.makedirs(final_output_dir, exist_ok=True)
    assets_dir = os.path.join(final_output_dir, 'assets')
    if os.path.lexists(assets_dir) and (
        os.path.islink(assets_dir) or not os.path.isdir(assets_dir)
    ):
        raise ValueError(f"assets 目录必须是普通目录（不允许符号链接）: {assets_dir}")
    os.makedirs(assets_dir, exist_ok=True)
    
    # 提取图片和Excel表格
    logger.info("正在提取内容...")
    skip_state = {}
    image_by_hash, table_queue_by_hash, table_repeat_by_hash = extract_content_from_docx(
        docx_path, assets_dir, on_limit=on_limit, skip_state=skip_state
    )
    skipped_resources = skip_state["skipped"]
    # Mammoth 回调必须复用 ZIP 提取阶段选定的同一份物理媒体
    # 白名单，不能按回调顺序重新分配配额，否则可落盘 2x 上限。
    allowed_media_paths = set(skip_state["allowed_media_paths"])
    media_quota_exceeded = any(
        "图片数量超过上限" in reason for _, reason in skipped_resources
    )
    noted_skipped = set()
    if any("图片数量超过上限" in reason for _, reason in skipped_resources):
        # 提取阶段已用一条记录汇总“剩余图片”；Mammoth 仍会
        # 逐引用回调，这里只返回占位符，不再按图片 hash 增长清单。
        noted_skipped.add("count")

    def _skipped_image_result(reason, image_data):
        """记录一次回调侧资源跳过，并返回可见跳过占位的 img src。"""
        marker = reason if reason == "count" else hashlib.sha256(image_data).hexdigest()[:8]
        if marker not in noted_skipped:
            noted_skipped.add(marker)
            skipped_resources.append(
                (f"文档引用图片#{marker}", SKIPPED_IMAGE_NOTE[reason]))
        return {"src": f"__SKIPPED_IMAGE_{reason}__"}

    table_md_by_placeholder = {}
    table_seq = [0]
    heading_level_map = extract_heading_level_map(docx_path)
    
    # 使用mammoth转换为HTML
    logger.info("正在转换文档...")

    def convert_image(image):
        """根据图片内容hash，返回对应的assets路径或表格占位符"""
        if skip_mode:
            media_key = _mammoth_embedded_media_key(image)
            if media_key is not None and media_key not in allowed_media_paths:
                return _skipped_image_result("count", b"")
            if media_key is None and media_quota_exceeded:
                # Mammoth 版本/图片类型无法暴露物理路径时选择安全回退：
                # 不在已确认超配额的文档中允许未知回调兜底落盘。
                return _skipped_image_result("count", b"")
            # skip 模式下回调读取同样有界，并复用提取阶段的大小/像素防线：
            # 已跳过的超限图片不会出现在 image_by_hash，若放行到下方兜底
            # 写盘分支即绕过防线，故超限在此直接返回可见跳过占位
            with image.open() as image_bytes:
                image_data = image_bytes.read(DOCX_SECURITY_LIMITS["image_file_size"] + 1)
            if len(image_data) > DOCX_SECURITY_LIMITS["image_file_size"]:
                return _skipped_image_result("size", image_data)
            pixels = image_pixel_count(image_data)
            if pixels is not None and pixels > DOCX_SECURITY_LIMITS["image_pixels"]:
                return _skipped_image_result("pixels", image_data)
            digest = hashlib.sha256(image_data).hexdigest()
        else:
            with image.open() as image_bytes:
                image_data = image_bytes.read()
            digest = hashlib.sha256(image_data).hexdigest()

        table_queue = table_queue_by_hash.get(digest)
        if table_queue:
            # 若仅剩一个元素则不再弹出，确保同一预览图多次出现时仍稳定替换为表格
            table_md = table_queue[0] if len(table_queue) == 1 else table_queue.pop(0)
            placeholder = f"__TABLE_PLACEHOLDER_{digest}_{table_seq[0]}__"
            table_seq[0] += 1
            table_md_by_placeholder[placeholder] = table_md
            return {"src": placeholder}

        # 防御性兜底：当前队列逻辑保证最后一个元素不会被弹出，因此此分支在
        # 正常流程中不会触发。保留此分支作为安全网，以防未来队列策略调整后
        # 队列被完全消耗的情况，确保仍能稳定替换为表格而非退化为普通图片。
        if digest in table_repeat_by_hash:
            table_md = table_repeat_by_hash[digest]
            placeholder = f"__TABLE_PLACEHOLDER_{digest}_{table_seq[0]}__"
            table_seq[0] += 1
            table_md_by_placeholder[placeholder] = table_md
            return {"src": placeholder}

        image_src = image_by_hash.get(digest)
        if image_src:
            return {"src": image_src}

        # 兜底：某些情况下zip里的图片与mammoth回调数据不一致，直接按hash写入assets
        ext = detect_image_format(image_data)
        if not ext:
            # 依据 mammoth 提供的 content_type 推断扩展名，仍未知则保留二进制原名
            content_subtype = (getattr(image, "content_type", "") or "").split("/")[-1].lower()
            ext = {
                "jpeg": ".jpeg", "jpg": ".jpeg", "png": ".png", "gif": ".gif",
                "webp": ".webp", "bmp": ".bmp", "tiff": ".tiff",
                "x-wmf": ".wmf", "x-emf": ".emf",
            }.get(content_subtype, ".bin")
        filename = f"image_{digest[:16]}{ext}"
        image_path = os.path.join(assets_dir, filename)
        if not os.path.exists(image_path):
            with open(image_path, "wb") as f:
                f.write(image_data)
        image_by_hash[digest] = f"assets/{filename}"
        return {"src": f"assets/{filename}"}
    
    with open(docx_path, 'rb') as docx_file:
        import mammoth

        result = mammoth.convert_to_html(
            docx_file,
            convert_image=mammoth.images.img_element(convert_image)
        )
        html = result.value
        for msg in getattr(result, "messages", []) or []:
            logger.debug("mammoth提示: %s", msg)
    
    # 将HTML转换为Markdown
    markdown = html_to_markdown(html, heading_level_map)
    
    # 替换表格占位符
    for placeholder_key, table_md in table_md_by_placeholder.items():
        placeholder = f"![]({placeholder_key})"
        markdown = markdown.replace(placeholder, f"\n\n{table_md}\n\n")

    # 自检：占位符未被替换通常意味着 mammoth 输出的 HTML 结构与预期不符，
    # 保留占位符并告警，便于发现未知文档结构的退化情况。
    leftover = re.findall(r"__TABLE_PLACEHOLDER_[0-9a-f]+_\d+__", markdown)
    if leftover:
        logger.warning("有 %d 个表格占位符未能替换为表格（请检查输出）", len(leftover))

    # 将 on_limit=skip 跳过的超限图片占位替换为可见说明
    # （不引用不存在的资源文件，跳过在输出中可见、可审计）
    markdown = re.sub(
        r"!\[[^\]]*\]\(__SKIPPED_IMAGE_([a-z]+)__\)",
        lambda m: f"*【图片已跳过：{SKIPPED_IMAGE_NOTE.get(m.group(1), m.group(1))}】*",
        markdown,
    )

    # 移除嵌入 Excel 替换后残留的预览图说明文本
    markdown = re.sub(
        r"\n+\*{0,2}点击图片可查看完整电子表格\*{0,2}\s*\n",
        "\n",
        markdown,
    )

    # 追加 mammoth 未能提取的文本框内容
    textbox_blocks = extract_textbox_content(docx_path)
    if textbox_blocks:
        # 检查主体中是否已包含文本框文本（mammoth 有时也能提取部分文本框）
        missing = [b for b in textbox_blocks if b.splitlines()[0] not in markdown]
        if missing:
            markdown += "\n\n---\n\n> **\\[文本框内容\\]**\n\n"
            for block in missing:
                markdown += f"> {block}\n>\n"
            logger.info("追加了 %d 个文本框内容", len(missing))

    # 追加 mammoth 未能提取的数学公式
    math_formulas = extract_math_text(docx_path)
    if math_formulas:
        missing_math = [f for f in math_formulas if f not in markdown]
        if missing_math:
            markdown += "\n\n---\n\n> **\\[数学公式\\]**\n\n"
            for formula in missing_math:
                markdown += f"> $$ {formula} $$\n>\n"
            logger.info("追加了 %d 个数学公式", len(missing_math))

    md_path = os.path.join(final_output_dir, f"{folder_name}.md")

    with open(md_path, 'w', encoding='utf-8') as f:
        f.write(markdown)

    # 子目录模式的 assets 归当前文档独占，可清理旧产物。
    # 平铺模式可能由多份 Markdown 共享 assets，不删除未引用文件。
    if create_subfolder:
        prune_stale_assets(assets_dir, image_by_hash.values())
    write_conversion_sentinel(
        final_output_dir, folder_name, source_sha256, on_limit=on_limit)

    if skip_mode and skipped_resources:
        logger.warning(
            "on_limit=skip：共跳过 %d 项超限资源（正文与其余资源已保留）", len(skipped_resources))
        for name, reason in skipped_resources:
            logger.warning("  已跳过: %s（%s）", name, reason)

    logger.info("转换完成: %s", md_path)
    return md_path


def _convert_footnotes(html: str) -> str:
    """将 mammoth 生成的脚注 HTML 转换为 Markdown 脚注语法。

    mammoth 输出格式：
      正文引用: <sup><a href="#footnote-N" id="footnote-ref-N">[N]</a></sup>
      文末列表: <li id="footnote-N"><p>text <a href="#footnote-ref-N">↑</a></p></li>
    """
    footnote_bodies: Dict[str, str] = {}

    def _extract_footnote_body(match):
        fid = match.group("fid")
        body_html = match.group("body")
        body_html = re.sub(r"</(?:p|div|li|br)\s*/?>", " ", body_html, flags=re.IGNORECASE)
        body = re.sub(r"<[^>]+>", "", body_html, flags=re.DOTALL)
        body = unescape(body).replace("↑", "").strip()
        body = re.sub(r"  +", " ", body)
        if body:
            footnote_bodies[fid] = body
        return ""

    html = re.sub(
        r'<li\b[^>]*\bid\s*=\s*["\']?footnote-(?P<fid>\d+)["\']?[^>]*>'
        r"(?P<body>.*?)</li>",
        _extract_footnote_body,
        html,
        flags=re.DOTALL | re.IGNORECASE,
    )

    html = re.sub(
        r"<sup>\s*<a\b[^>]*href\s*=\s*[\"']?#footnote-(\d+)[\"']?[^>]*>"
        r"\s*\[\d+\]\s*</a>\s*</sup>",
        lambda m: f"[^{m.group(1)}]",
        html,
        flags=re.IGNORECASE,
    )

    if footnote_bodies:
        footer = "\n\n---\n\n"
        for fid in sorted(footnote_bodies, key=int):
            footer += f"[^{fid}]: {footnote_bodies[fid]}\n"
        html += footer

    return html


def html_to_markdown(html, heading_level_map: Optional[Dict[str, int]] = None):
    """将HTML转换为Markdown"""

    html = _convert_footnotes(html)

    # 处理标题
    html = re.sub(r'<h1[^>]*>(.*?)</h1>', r'# \1\n\n', html, flags=re.DOTALL)
    html = re.sub(r'<h2[^>]*>(.*?)</h2>', r'## \1\n\n', html, flags=re.DOTALL)
    html = re.sub(r'<h3[^>]*>(.*?)</h3>', r'### \1\n\n', html, flags=re.DOTALL)
    html = re.sub(r'<h4[^>]*>(.*?)</h4>', r'#### \1\n\n', html, flags=re.DOTALL)
    html = re.sub(r'<h5[^>]*>(.*?)</h5>', r'##### \1\n\n', html, flags=re.DOTALL)
    html = re.sub(r'<h6[^>]*>(.*?)</h6>', r'###### \1\n\n', html, flags=re.DOTALL)
    
    # 优先按 DOCX 原始 heading 样式提升标题层级（主流程），避免纯文本启发式误判。
    if heading_level_map:
        def _replace_anchored_heading(match):
            heading_id = match.group("id1") or match.group("id2") or match.group("id3") or ""
            content = match.group("content")
            level = heading_level_map.get(heading_id)
            if not level:
                return match.group(0)
            text = re.sub(r"<[^>]+>", "", content, flags=re.DOTALL)
            text = unescape(text).strip()
            if not text:
                return ""
            return f"{'#' * level} {text}\n\n"

        html = re.sub(
            (
                r"<p[^>]*>\s*"
                r"<a[^>]*\bid\s*=\s*(?:\"(?P<id1>heading_\d+)\"|'(?P<id2>heading_\d+)'|(?P<id3>heading_\d+))[^>]*>"
                r"\s*</a>(?P<content>.*?)</p>"
            ),
            _replace_anchored_heading,
            html,
            flags=re.DOTALL | re.IGNORECASE,
        )

    # 处理粗体和斜体
    html = re.sub(r'<strong>(.*?)</strong>', r'**\1**', html, flags=re.DOTALL)
    html = re.sub(r'<b>(.*?)</b>', r'**\1**', html, flags=re.DOTALL)
    html = re.sub(r'<em>(.*?)</em>', r'*\1*', html, flags=re.DOTALL)
    html = re.sub(r'<i>(.*?)</i>', r'*\1*', html, flags=re.DOTALL)
    
    # 处理图片
    def _replace_img(match):
        src = match.group("src1") or match.group("src2") or match.group("src3") or ""
        return f"![]({src})\n\n"

    html = re.sub(
        (
            r"<img\b[^>]*\bsrc\s*=\s*"
            r"(?:\"(?P<src1>[^\"]*)\"|'(?P<src2>[^']*)'|(?P<src3>[^\s\"'=<>`]+))"
            r"[^>]*/?>"
        ),
        _replace_img,
        html,
        flags=re.IGNORECASE,
    )
    
    # 处理链接（支持双引号、单引号、无引号三种 href 写法）
    def _replace_link(match):
        href = match.group("href1") or match.group("href2") or match.group("href3") or ""
        text = match.group("text")
        return f"[{text}]({href})"

    html = re.sub(
        (
            r"<a\b[^>]*\bhref\s*=\s*"
            r"(?:\"(?P<href1>[^\"]*)\"|'(?P<href2>[^']*)'|(?P<href3>[^\s\"'=<>`]+))"
            r"[^>]*>(?P<text>.*?)</a>"
        ),
        _replace_link,
        html,
        flags=re.DOTALL | re.IGNORECASE,
    )

    # 先把HTML里的换行标签转为文本换行（需早于表格转换，避免改写表格里的 <br> 文本）
    html = re.sub(r'<br\s*/?>', '\n', html)

    # 先处理表格（必须在段落/列表转换之前）
    html = re.sub(
        r'<table[^>]*>.*?</table>',
        lambda match: table_html_to_markdown(match.group(0)),
        html,
        flags=re.DOTALL | re.IGNORECASE,
    )

    # 使用结构化解析处理列表，避免正则顺序导致的嵌套层级破坏。
    html = transform_html_lists_to_markdown(html)
    
    # 处理段落
    html = re.sub(r'<p[^>]*>(.*?)</p>', r'\1\n\n', html, flags=re.DOTALL)
    
    # 移除剩余的HTML标签（保留 <br> 供 Markdown 单元格换行显示）
    html = re.sub(r'<(?!br\s*/?)[^>]+>', '', html, flags=re.IGNORECASE)
    
    # 清理多余的空行
    html = re.sub(r'\n{3,}', '\n\n', html)
    
    html = html.replace('&nbsp;', ' ')
    html = unescape(html)

    html = promote_numbered_bold_headings(html)
    html = promote_leading_bold_title(html)
    return html.strip()


if __name__ == '__main__':
    import argparse

    logging.basicConfig(level=logging.INFO, format="%(message)s")

    parser = argparse.ArgumentParser(
        description="将单个 DOCX 文档转换为 Markdown（提取图片与嵌入 Excel 表格）")
    parser.add_argument("docx_path", help="DOCX 文件路径")
    parser.add_argument("output_dir", help="输出目录路径")
    parser.add_argument("--output-name", default=None,
                        help="自定义输出子文件夹与 .md 文件名（末尾 .docx 自动去除；默认用源文件名）")
    parser.add_argument("--on-limit", choices=("reject", "skip"), default="reject",
                        help="资源超限处置：reject 整篇拒绝（默认）；skip 仅跳过超限资源继续转换"
                             "（ZIP bomb 等恶意特征仍整篇拒绝）")
    args = parser.parse_args()

    docx_path = args.docx_path

    if not os.path.exists(docx_path):
        logger.error("文件不存在 - %s", docx_path)
        sys.exit(1)

    try:
        convert_docx_to_markdown(
            docx_path, args.output_dir,
            output_name=args.output_name, on_limit=args.on_limit)
    except DocxSecurityError as exc:
        logger.error("安全拒绝（输入恶意或资源超限，不重试）: %s", exc)
        sys.exit(2)
    except ValueError as exc:
        logger.error("输入错误: %s", exc)
        sys.exit(2)
