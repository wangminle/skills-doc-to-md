import importlib.util
import io
import tempfile
import unittest
import zipfile
from pathlib import Path


WORKSPACE_ROOT = Path(__file__).resolve().parents[1]
SCRIPT_PATH = WORKSPACE_ROOT / "skills" / "docx-to-markdown" / "scripts" / "convert_docx.py"
DOCX_PATH = WORKSPACE_ROOT / "tests" / "自研语义VAD（云端VAD-4.0）型号接入需求文档V2.docx"
DOCX_DIST_WAKE = WORKSPACE_ROOT / "tests" / "分布式唤醒V1.9.1—【唤醒体验v03】播控指令与唤醒暂停的策略梳理.docx"


def load_convert_module():
    spec = importlib.util.spec_from_file_location("convert_docx_module", SCRIPT_PATH)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


class TestEmbeddedExcelTables(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.convert = load_convert_module()

    @unittest.skipUnless(DOCX_PATH.is_file(), "可选真实 DOCX 回归夹具未提供")
    def test_excel_sheet2_can_be_converted_to_markdown(self):
        with zipfile.ZipFile(DOCX_PATH, "r") as zf:
            xlsx_data = zf.read("word/embeddings/Microsoft_Excel_Worksheet2.xlsx")

        markdown = self.convert.excel_to_markdown(xlsx_data)
        self.assertIsNotNone(markdown)
        self.assertIn("单意图截断准确率", markdown)
        self.assertIn("讯飞语义VAD（VAD-3.0）", markdown)
        self.assertIn("98.3%↗", markdown)

    @unittest.skipUnless(DOCX_PATH.is_file(), "可选真实 DOCX 回归夹具未提供")
    def test_docx_conversion_contains_two_expected_tables(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            md_path = self.convert.convert_docx_to_markdown(str(DOCX_PATH), tmpdir)
            content = Path(md_path).read_text(encoding="utf-8")

        self.assertIn("| 文档版本 | 修改内容 | 修改日期 | 作者 |", content)
        self.assertIn("| 讯飞语义VAD（VAD-3.0） | 0.945 | 0.883 | 2.203s |", content)
        self.assertIn("| 自研语义VAD（VAD-4.0） | 98.3%↗ | 93.6%↗ | 1.797s↗ |", content)

    @unittest.skipUnless(DOCX_DIST_WAKE.is_file(), "可选真实 DOCX 回归夹具未提供")
    def test_excel_merged_cells_are_expanded_to_each_body_row(self):
        with zipfile.ZipFile(DOCX_DIST_WAKE, "r") as zf:
            xlsx_data = zf.read("word/embeddings/Microsoft_Excel_Worksheet1.xlsx")

        markdown = self.convert.excel_to_markdown(xlsx_data)
        self.assertIsNotNone(markdown)
        self.assertIn("| 唤醒设备 | / | 播放TTS话术 | 打断 | 应答 |", markdown)
        self.assertIn("| 唤醒设备 | / | 无动作 | 无 | 应答 |", markdown)
        self.assertIn("| 非唤醒设备 | 不同房间 | 播放TTS话术 | 打断 | 无 |", markdown)

    def test_excel_merged_cells_expand_vertical_horizontal_and_block(self):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        ws["A1"] = "V"
        ws["B1"] = "H"
        ws["A2"] = "R1"
        ws["B2"] = "R2"
        ws["C2"] = "R3"

        # 纵向合并
        ws.merge_cells("A1:A3")
        # 横向合并
        ws.merge_cells("B1:D1")
        # 矩形合并 (2x2)
        ws["E2"] = "M"
        ws.merge_cells("E2:F3")

        buf = io.BytesIO()
        wb.save(buf)
        wb.close()

        markdown = self.convert.excel_to_markdown(buf.getvalue())
        self.assertIsNotNone(markdown)
        lines = [line for line in markdown.splitlines() if line.startswith("|")]

        self.assertIn("> merge_ranges:", markdown)
        self.assertIn("A1:A3", markdown)
        self.assertIn("B1:D1", markdown)
        self.assertIn("E2:F3", markdown)

        # 表头：横向合并被展开
        self.assertIn("| V | H | H | H |", lines[0])
        # 第二行：纵向值和矩形值均存在
        self.assertIn("| V | R2 | R3 |  | M | M |", markdown)
        # 第三行：纵向与矩形继续展开
        self.assertIn("| V |  |  |  | M | M |", markdown)

    def test_excel_prefers_data_sheet_over_large_merged_template_sheet(self):
        import openpyxl

        wb = openpyxl.Workbook()
        ws_template = wb.active
        ws_template.title = "Template"
        ws_template["A1"] = "封面标题"
        ws_template.merge_cells("A1:Z100")

        ws_data = wb.create_sheet("Data")
        ws_data.append(["col1", "col2", "col3"])
        ws_data.append(["1-1", "1-2", "1-3"])
        ws_data.append(["2-1", "2-2", "2-3"])

        buf = io.BytesIO()
        wb.save(buf)
        wb.close()

        markdown = self.convert.excel_to_markdown(buf.getvalue())
        self.assertIsNotNone(markdown)
        self.assertIn("| col1 | col2 | col3 |", markdown)
        self.assertIn("| 1-1 | 1-2 | 1-3 |", markdown)
        self.assertNotIn("封面标题", markdown)

    def test_format_cell_value_handles_non_finite_floats(self):
        self.assertEqual(self.convert._format_cell_value(float("nan")), "nan")
        self.assertEqual(self.convert._format_cell_value(float("inf")), "inf")
        self.assertEqual(self.convert._format_cell_value(float("-inf")), "-inf")

if __name__ == "__main__":
    unittest.main()
